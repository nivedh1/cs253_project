from django.shortcuts import render,redirect
from app1.models import Profile,Student
from django.contrib.auth.models import User, auth
#import . from models
import openpyxl


def upload(request):
    if "GET" == request.method:
        return render(request, 'upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()

        for row in worksheet.iter_rows():
            print(int(row[0].value))
            print(str(row[1].value))
            p=Student.objects.create(number   =  int(row[0].value))

            p.first_name = str(row[1].value)
            p.last_name = str(row[2].value)
            p.department = str(row[3].value)
            p.save()
            print(p.number) 


        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'upload.html', {"excel_data":excel_data})



def register(request):
    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']

        user_dummy = User.objects.create_user(username=username,password=password1,email=email,first_name=first_name,last_name=last_name)
        user_dummy.save()
        #role=request.POST['role']
        pro=Profile.objects.create(user = user_dummy,role='Ta')
        
        pro.start_No=41
        pro.end_No=44

        #user=user_dummy
        
        pro.save()
        print("hi")

        print(pro.role)
        print(pro.user.username)

        #add messages latter
        return redirect('/login')
    else:
        return render(request,'register.html')


def login(request):
    if(request.method=="POST"):
        print("check_2")
        username=request.POST['username']
        password=request.POST['password']

        user=auth.authenticate(username=username,password=password)

        if user is not None:
            auth.login(request,user)
            actual_user = Profile.objects.get(user=user)
            i=0
            if(actual_user.role=='Ta'):
                i=i+1
                #return render(request,'ta_dashboard.html',{'user':actual_user},)

                #code for ta
            
            elif(actual_user.role=='Manager'):
                i=i+1
                #code for manager
            else:
                i=i+1
                #code for admin

            print("ta login successful")
            
            arr = list()
            for num in range(actual_user.start_No,actual_user.end_No+1):
                stu = Student.objects.get(number=num)
                print(stu.first_name)
                arr.append(stu)
            print(arr[0].first_name)

            return render(request,'ta_dashboard.html',{'user':actual_user,"arr":arr})
        else:
            print('invalid Credentials')
            return render(request,'login.html')
    else:
        print("check4")
        return render(request,'login.html')


def student(request,student_id):
    if(request.method=="GET"):
        user=auth.get_user(request)
        if(user.is_authenticated):
            actual_user = Profile.objects.get(user=user)
            if(actual_user.start_No<=student_id or actual_user.end_No>=student_id):
                stu=Student.objects.get(number=student_id)
                return render(request,'student.html',{'user':actual_user,'student':stu})
            else:
                return redirect('/login')
                #this redirection supposed to be changed 

        #    return redirect('/login')
        #print("jii")
        else:
            return redirect('/login')


    else:
        discrepency=request.POST.get('discrepency')
        
        comment=request.POST['comment']
        stu=Student.objects.get(number=student_id)
        stu.check_box='Yes'
        print(stu.discrepency)
        if(str(discrepency)=='on'):
            stu.discrepency='Yes'
        else:
            stu.discrepency='No'


        print(stu.discrepency)
        stu.comment=comment
        stu.save()
        user=auth.get_user(request)
        actual_user = Profile.objects.get(user=user)
        arr=list()
        for num in range(actual_user.start_No,actual_user.end_No+1):
            student = Student.objects.get(number=num)
            print(student.first_name)
            arr.append(student)

        return render(request,'ta_dashboard.html',{'user':actual_user,"arr":arr})



    








def logout(request):
    print(auth.get_user(request).username)

    print("logout working")
    auth.logout(request)
    print(auth.get_user(request).username)
    if(auth.get_user(request).is_authenticated):
        print("hii")
    else:
        print("bye")
    return redirect('/login')

